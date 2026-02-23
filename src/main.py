import time
import pathlib
import sys
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import openpyxl

# ==========================================
# CONFIGURATION & CONSTANTS
# ==========================================

# Define the path to the source file being monitored.
# ideally, place this file in the same directory as the script or update the path.
SOURCE_FILE_NAME = "source_master_data.xlsx"
SOURCE_PATH = pathlib.Path(SOURCE_FILE_NAME).resolve()

# --- MAPPING TEMPLATES ---
# Define which cells in the Source map to which cells in the Target.
# Structure: { "Source_Cell": "Target_Cell" } or { "Source_Cell": ["Target_Cell_1", "Target_Cell_2"] }

# [Sample] Mapping for General Information
general_info_mapping = {
    "B1": "C3",  # Example: Album Title -> Target Cell
    "B2": "G3",  # Example: Artist Name -> Target Cell
    "B3": "C5",  # Example: Release Date -> Target Cell
    # Add your specific cell mappings here...
}

# [Sample] Table Mapping (for list-based data like tracklists)
# Structure: { "Source_Column_Index": Target_Column_Index }
# Note: Column indices are 1-based (A=1, B=2, etc.)
tracklist_table_mapping = {
    "1": 1,    # Disk Number
    "2": 2,    # Track Number
    "3": 12,   # ISRC Code
    "6": 5,    # Track Title
    "8": None, # Ignore this column
    # Add your specific column mappings here...
}

# --- TARGET CONFIGURATION ---
# List of dictionaries defining each target file and its specific rules.
TARGETS = [
    {
        "id": "target_a",
        "filename": "target_report_A.xlsx",
        "sheet_name": "Sheet1",
        "row_offset": 5,  # Row difference between source table and target table
        "mapping": general_info_mapping,
        "table_mapping": tracklist_table_mapping
    },
    {
        "id": "target_b",
        "filename": "target_platform_upload.xlsx",
        "sheet_name": "TrackList",
        "row_offset": -2, 
        "mapping": general_info_mapping, # Re-using mapping or define a new one
        "table_mapping": tracklist_table_mapping
    },
    # Add more target configurations as needed
]


# ==========================================
# CORE LOGIC
# ==========================================

def sync_data():
    """
    Reads data from the source workbook and updates all target workbooks
    defined in the configuration.
    """
    print(f"[{time.strftime('%H:%M:%S')}] Detected change. Starting sync...")

    # 1. Load Source Workbook
    try:
        # data_only=True ensures we get the calculated values, not formulas
        wb_src = openpyxl.load_workbook(SOURCE_PATH, data_only=True)
        ws_src = wb_src.active
    except PermissionError:
        print(f"[Error] Could not open source file. Please close '{SOURCE_FILE_NAME}' and try again.")
        return
    except FileNotFoundError:
        print(f"[Error] Source file '{SOURCE_FILE_NAME}' not found.")
        return
    except Exception as e:
        print(f"[Error] Unexpected error opening source: {e}")
        return

    # 2. Iterate through Targets
    for target_config in TARGETS:
        target_path = pathlib.Path(target_config["filename"]).resolve()

        try:
            wb_target = openpyxl.load_workbook(target_path)
        except FileNotFoundError:
            print(f"[Warning] Target file '{target_config['filename']}' not found. Skipping.")
            continue
        except PermissionError:
            print(f"[Skipped] Target file '{target_config['filename']}' is currently open by a user.")
            continue
        
        # Check if sheet exists
        if target_config["sheet_name"] not in wb_target.sheetnames:
            print(f"[Warning] Sheet '{target_config['sheet_name']}' not found in '{target_config['filename']}'. Skipping.")
            wb_target.close()
            continue

        ws_target = wb_target[target_config["sheet_name"]]

        # --- A. Apply Single Cell Mappings ---
        for src_addr, target_addr in target_config["mapping"].items():
            val = ws_src[src_addr].value
            
            # Handle one-to-many mapping
            if isinstance(target_addr, list):
                for addr in target_addr:
                    ws_target[addr].value = val
            else:
                ws_target[target_addr].value = val

        # --- B. Apply Dynamic Table/Row Mappings ---
        # Assuming table data starts at row 13 in source (as per original logic)
        START_ROW = 13 
        row_offset = target_config["row_offset"]
        
        for src_col, target_col in target_config["table_mapping"].items():
            if target_col is None:
                continue

            for row in range(START_ROW, ws_src.max_row + 1):
                cell_val = ws_src.cell(row=row, column=int(src_col)).value
                
                # Calculate target row
                target_row = row + row_offset
                
                if isinstance(target_col, list):
                    for t_col in target_col:
                        ws_target.cell(row=target_row, column=t_col).value = cell_val
                else:
                    ws_target.cell(row=target_row, column=target_col).value = cell_val

        # --- C. Custom Logic (Example: Notice Generation) ---
        # If specific formatting logic is needed for a specific target ID
        if target_config["id"] == "target_a":
            # Example: Concatenate generic info into a header
            date_val = str(ws_src["D6"].value or "")
            ws_target["A1"].value = f"[Auto-Generated] Release Notice - {date_val}"

        # 3. Save Target Workbook
        try:
            wb_target.save(target_path)
            print(f"   -> Synced: {target_config['filename']}")
        except PermissionError:
            print(f"   -> [Failed] Could not save '{target_config['filename']}'. Is it open?")
        except Exception as e:
            print(f"   -> [Error] Saving '{target_config['filename']}': {e}")
        finally:
            wb_target.close()

    # Close source to free resources
    wb_src.close()
    print(f"[{time.strftime('%H:%M:%S')}] Sync complete.\n")


class ExcelChangeHandler(FileSystemEventHandler):
    """
    Handles file system events. Includes a debounce mechanism to prevent 
    duplicate triggers when Excel performs multiple write operations during a save.
    """
    def __init__(self):
        self.last_modified = 0
        self.debounce_interval = 2.0  # Seconds to wait before allowing a new sync

    def on_modified(self, event):
        # We only care about the specific source file
        if pathlib.Path(event.src_path).resolve() == SOURCE_PATH:
            current_time = time.time()
            if (current_time - self.last_modified) > self.debounce_interval:
                self.last_modified = current_time
                # Small sleep to ensure Excel releases the file lock
                time.sleep(0.5)
                sync_data()

if __name__ == "__main__":
    # Ensure source path exists before starting
    if not SOURCE_PATH.exists():
        print(f"[Error] Source file '{SOURCE_FILE_NAME}' does not exist in the current directory.")
        print(f"Expected path: {SOURCE_PATH}")
        sys.exit(1)

    event_handler = ExcelChangeHandler()
    observer = Observer()
    
    # Watch the directory containing the source file
    observer.schedule(event_handler, path=str(SOURCE_PATH.parent), recursive=False)
    observer.start()

    print(f"==================================================")
    print(f" Excel Auto-Sync Tool Running")
    print(f" Monitoring: {SOURCE_FILE_NAME}")
    print(f" Press Ctrl+C to stop")
    print(f"==================================================\n")

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
        print("\nStopping observer...")
    
    observer.join()